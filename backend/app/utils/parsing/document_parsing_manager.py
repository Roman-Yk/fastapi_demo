import io
import os
import docx2txt
import openpyxl
import PyPDF2
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image
from typing import List
from app.core.configs.FileConfig import FileConfig


class DocumentParsingManager:
    """
    A class to handle parsing of various document types and extracting text from them.
    This class provides methods to extract text from PDF, image, Excel, Word, and text files.
    It uses libraries like PyPDF2, pytesseract, openpyxl, and docx2txt for parsing.
    """

    CUSTOM_TESSERACT_CONFIG = r"--oem 1 --psm 11"
    PARSING_LANGUAGES = "nor+eng"

    @classmethod
    def get_document_text_based_on_file_extension(cls, file_bytes, extension: str) -> str:
        """
        Extracts text from a document based on its file extension.
        Args:
                file_bytes (bytes or io.BytesIO): The document file in bytes or a BytesIO object.
                extension (str): The file extension of the document.
        Returns:
                str: The extracted text from the document.
        """
        if extension == ".pdf":
            text = cls.get_text_from_pdf_with_pypdf2(file_bytes)
            text += cls.get_text_from_pdf_with_pytesseract(file_bytes)
        elif extension in FileConfig.images_extensions:
            text = cls.get_text_from_image_with_pytesseract(file_bytes)
        elif extension == ".xlsx":
            text = cls.get_text_from_xlsx_with_openpyxl(file_bytes)
        elif extension == ".docx":
            text = cls.get_text_from_docx_with_docx2txt(file_bytes)
        elif extension == ".txt":
            text = cls.get_text_from_txt_with_default(file_bytes)
        else:
            text = ""
        return text


    @classmethod
    def get_file_extension_from_path(cls, file_path: str):
        """
        Extracts the file extension from a given file path.
        Args:
                file_path (str): The path to the file.
        Returns:
                str: The file extension in lowercase.
        """
        return os.path.splitext(file_path)[-1].lower()


    @classmethod
    def get_text_from_document_by_path(cls, file_path: str) -> str:
        """
        Reads the content of a document file and returns it as bytes.
        Args:
                file_path (str): The path to the document file.
        Returns:
                str: The content of the document file in bytes.
        """
        extension = os.path.splitext(file_path)[-1].lower()

        with open(file_path, "rb") as f:
            file_bytes = f.read()

        return DocumentParsingManager.get_document_text_based_on_file_extension(
            file_bytes, extension
        )


    @classmethod
    def get_text_from_image_with_pytesseract(cls, file_bytes) -> str:
        """
        Extracts text from an image using pytesseract.
        Args:
                file_bytes (bytes or io.BytesIO): The image file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the image.
        """
        if isinstance(file_bytes, bytes):
            file_bytes = io.BytesIO(file_bytes)

        with Image.open(file_bytes) as image:
            if image.mode != "RGB":
                image = image.convert("RGB")

            max_dim = 1600
            if image.width > max_dim or image.height > max_dim:
                scale = max_dim / max(image.width, image.height)
                image = image.resize(
                    (int(image.width * scale), int(image.height * scale))
                )

            psm_modes = ["11", "6", "3"]
            for psm in psm_modes:
                try:
                    config = f"--oem 3 --psm {psm}"
                    return pytesseract.image_to_string(
                        image, lang=cls.PARSING_LANGUAGES, config=config
                    )
                except Exception as error:
                    continue
            return pytesseract.image_to_string(
                image, lang=cls.PARSING_LANGUAGES, config=cls.CUSTOM_TESSERACT_CONFIG
            )


    @classmethod
    def get_text_from_pdf_with_pytesseract(cls, file_bytes) -> str:
        """
        Extracts text from a PDF file using pytesseract.
        Args:
                file_bytes (bytes or io.BytesIO): The PDF file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the PDF.
        """
        # Handle both bytes and BytesIO objects
        if isinstance(file_bytes, bytes):
            pdf_bytes = file_bytes
        else:
            pdf_bytes = file_bytes.getvalue()

        doc = convert_from_bytes(pdf_bytes)
        text_list = []
        for _, page_data in enumerate(doc):
            image_string = pytesseract.image_to_string(
                page_data,
                lang=cls.PARSING_LANGUAGES,
                config=cls.CUSTOM_TESSERACT_CONFIG,
            )
            text_list.append(f"\n{image_string}")
        return "".join(text_list)


    @classmethod
    def get_text_from_pdf_with_pypdf2(cls, file_bytes) -> str:
        """
        Extracts text from a PDF file using PyPDF2.
        Args:
                file_bytes (bytes or io.BytesIO): The PDF file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the PDF.
        """
        if isinstance(file_bytes, bytes):
            file_bytes = io.BytesIO(file_bytes)

        try:
            pdf_reader = PyPDF2.PdfReader(file_bytes)
            text_list = []

            for i in range(len(pdf_reader.pages)):
                page_text = pdf_reader.pages[i].extract_text()
                text_list.append(f"\n{page_text}")

            return "".join(text_list)
        except Exception as e:
            print(f"Error parsing PDF with PyPDF2: {e}")
            return ""


    @classmethod
    def _get_column_widths(cls, worksheet) -> List:
        """
        Calculates the maximum width of each column in the worksheet.
        Args:
                worksheet: The worksheet object from openpyxl.
        Returns:
                list: A list of maximum widths for each column.
        """
        column_widths = []
        has_2_rows = worksheet.max_row >= 2
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if has_2_rows:
                column_widths.append(
                    max(len(str(col[0].value)), len(str(col[1].value)))
                )
            else:
                column_widths.append(len(str(col[0].value)))
        return column_widths


    @classmethod
    def get_text_from_worksheet(cls, worksheet) -> str:
        """
        Extracts text from an openpyxl worksheet and formats it into a string.
        Args:
                worksheet: The worksheet object from openpyxl.
        Returns:
                str: The formatted text from the worksheet.
        """
        text_list = []
        column_widths = cls._get_column_widths(worksheet)
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                width = column_widths[col[i].column - 1]
                text_list.append(f"{str(col[i].value):{width}}   |")
            text_list.append("\n")
        return "".join(text_list)


    @classmethod
    def get_text_from_xlsx_with_openpyxl(cls, file_bytes) -> str:
        """
        Extracts text from an Excel file using openpyxl.
        Args:
                file_bytes (bytes or io.BytesIO): The Excel file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the Excel file.
        """
        workbook = openpyxl.load_workbook(file_bytes)
        sheet_names = workbook.sheetnames
        text_list = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            text_list.append(f"{worksheet}\n{cls.get_text_from_worksheet(worksheet)}\n")
        return "".join(text_list)


    @classmethod
    def get_text_from_docx_with_docx2txt(cls, file_bytes) -> str:
        """
        Extracts text from a DOCX file using docx2txt.
        Args:
                file_bytes (bytes or io.BytesIO): The DOCX file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the DOCX file.
        """
        return docx2txt.process(file_bytes)


    @classmethod
    def get_text_from_txt_with_default(cls, file_bytes) -> str:
        """
        Extracts text from a TXT file using the default read method.
        Args:
                file_bytes (bytes or io.BytesIO): The TXT file in bytes or a BytesIO object.
        Returns:
                str: The extracted text from the TXT file.
        """
        file_bytes.seek(0)
        data = file_bytes.read()
        return data.decode("utf-8")
